import os
import json
import uuid
import logging
import hashlib
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
import traceback
import time
import subprocess

import pandas as pd
import pyodbc
import dbf
import sqlite3
import numpy as np
from mongoengine import get_db

from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.response import Response
from rest_framework import status

from pymongo import MongoClient, errors, UpdateOne

logger = logging.getLogger(__name__)

# ================== MONGO CONFIG ==================
MONGO_HOST = os.getenv("MONGO_HOST", "localhost")
MONGO_PORT = int(os.getenv("MONGO_PORT", 27017))

MONGO_URI = f"mongodb://{MONGO_HOST}:{MONGO_PORT}/?directConnection=true"

# ---- Database & Collection names ----
MAPPING_DB_NAME = "Watchlist"
MAPPING_COLLECTION_NAME = os.getenv("MAPPING_COLLECTION", "Watchlist_cols")

WATCHLIST_DB_NAME = "Watchlist"
WATCHLIST_COLLECTION_NAME = "WatchList_data"
WATCHLIST_NEXUS_COLLECTION_NAME = "Watchlist_nexus"

# ---- Global Mongo Client (reuse everywhere) ----
mongo_client = MongoClient(
    MONGO_URI,
    serverSelectionTimeoutMS=5000,
    maxPoolSize=50
)

# ---- DB handles ----
MAPPING_DB = mongo_client[MAPPING_DB_NAME]
WATCHLIST_DB = mongo_client[WATCHLIST_DB_NAME]

# ---- Collection handles ----
MAPPING_COLLECTION = MAPPING_DB[MAPPING_COLLECTION_NAME]
WATCHLIST_COLLECTION = WATCHLIST_DB[WATCHLIST_COLLECTION_NAME]
WATCHLIST_NEXUS_COLLECTION = WATCHLIST_DB[WATCHLIST_NEXUS_COLLECTION_NAME]

TEMP_DIR = "/tmp/subscriber_uploads"
os.makedirs(TEMP_DIR, exist_ok=True)

ALLOWED_EXTENSIONS = {
    'txt', 'csv', 'xls', 'xlsx',
    'mdb', 'accdb', 'dbf',
    'db', 'sqlite', 'sqlite3'
}

# ================== PERFORMANCE CONFIG ==================
CHUNK_SIZE = 5000
MAX_WORKERS = 8
BATCH_SIZE = 1000
PREVIEW_ROWS = 10

# Track background save status
background_save_status = {}
background_save_lock = Lock()

# Track import progress
import_progress = {}
import_progress_lock = Lock()


# ================= HELPER FUNCTIONS =================
def generate_hash_id(group_name, sub_group_name):
    """
    ✅ Generate 16-character hash ID from GroupName and Sub_Group_Name
    """
    combined = f"{group_name}_{sub_group_name}"
    hash_obj = hashlib.md5(combined.encode())
    return hash_obj.hexdigest()[:16]


# ================= FILE READER ==============
class FileColumnReader:
    """
    ✅ FIXED: Now properly handles multi-table/sheet files:
    - Excel: Multiple sheets
    - SQLite: Multiple tables
    - Access: Multiple tables (Docker compatible with mdb-tools)
    """

    def __init__(self, file_object, filename):
        self.file_object = file_object
        self.filename = filename
        self.ext = Path(filename).suffix.lower()

    def has_multiple_tables(self):
        """Check if file type can have multiple tables/sheets"""
        return self.ext in ['.xls', '.xlsx', '.db', '.sqlite', '.sqlite3', '.mdb', '.accdb']

    def get_all_tables_or_sheets(self):
        """
        ✅ NEW: Get list of all tables/sheets in the file

        Returns:
            dict or None: {
                'type': 'excel|sqlite|access',
                'tables': ['Sheet1', 'Sheet2', ...],
                'count': 3
            }
        """
        try:
            if self.ext in ['.xls', '.xlsx']:
                sheets = self.get_excel_sheets()
                if sheets:
                    return {'type': 'excel', 'tables': sheets, 'count': len(sheets)}
            elif self.ext in ['.db', '.sqlite', '.sqlite3']:
                tables = self.get_sqlite_tables()
                if tables:
                    return {'type': 'sqlite', 'tables': tables, 'count': len(tables)}
            elif self.ext in ['.mdb', '.accdb']:
                tables = self.get_access_tables()
                if tables:
                    return {'type': 'access', 'tables': tables, 'count': len(tables)}
            return None
        except Exception as e:
            logger.error(f"Error getting tables/sheets: {str(e)}")
            return None

    def read_data(self, table_name=None, limit=None, sheet_name=None):
        """
        ✅ ENHANCED: Now supports sheet_name parameter for Excel files

        Args:
            table_name: For database files (SQLite, Access)
            limit: Max rows to read
            sheet_name: ✅ NEW - For Excel files, specific sheet to read
        """
        try:
            if self.ext == '.csv':
                return self.read_csv(limit)

            if self.ext in ['.xls', '.xlsx']:
                return self.read_excel(limit, sheet_name=sheet_name)

            if self.ext == '.txt':
                return self.read_txt(limit)

            if self.ext == '.dbf':
                return self.read_dbf(limit)

            if self.ext in ['.db', '.sqlite', '.sqlite3']:
                return self.read_sqlite(table_name, limit)

            if self.ext in ['.mdb', '.accdb']:
                return self.read_access(table_name, limit)

            raise Exception(f"Unsupported file type: {self.ext}")

        except Exception as e:
            raise Exception(f"Failed to read file: {str(e)}")

    def read_csv(self, limit):
        """Handle malformed CSV files"""
        self.file_object.seek(0)
        try:
            df = pd.read_csv(
                self.file_object,
                nrows=limit,
                on_bad_lines='skip',
                low_memory=False
            )
        except TypeError:
            df = pd.read_csv(
                self.file_object,
                nrows=limit,
                error_bad_lines=False,
                warn_bad_lines=True,
                low_memory=False
            )
        return self._normalize_columns(df)

    def read_excel(self, limit, sheet_name=None):
        """
        ✅ FIXED: Now supports sheet selection

        Args:
            limit: Max rows to read
            sheet_name: Specific sheet to read. If None, reads first sheet.
        """
        self.file_object.seek(0)

        try:
            # Get all available sheets
            xl_file = pd.ExcelFile(self.file_object)
            available_sheets = xl_file.sheet_names

            logger.info(f"📊 Excel file has {len(available_sheets)} sheet(s): {available_sheets}")

            # Determine which sheet to read
            if sheet_name:
                if sheet_name not in available_sheets:
                    xl_file.close()
                    raise Exception(
                        f"Sheet '{sheet_name}' not found. "
                        f"Available sheets: {available_sheets}"
                    )
                target_sheet = sheet_name
                logger.info(f"📖 Reading selected sheet: '{sheet_name}'")
            else:
                target_sheet = 0  # First sheet (default behavior - backward compatible)
                logger.info(f"📖 Reading first sheet: '{available_sheets[0]}'")

            # Read the sheet
            df = pd.read_excel(xl_file, sheet_name=target_sheet, nrows=limit)
            xl_file.close()

            logger.info(f"✅ Successfully read {len(df)} rows from Excel sheet")
            return self._normalize_columns(df)

        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            raise

    def read_txt(self, limit):
        """Handle malformed TXT files"""
        self.file_object.seek(0)
        first = self.file_object.readline().decode().strip()
        self.file_object.seek(0)

        for d in ['\t', ',', '|', ';']:
            if len(first.split(d)) > 1:
                try:
                    logger.info(f"📖 Reading with delimiter: '{d}'")
                    df = pd.read_csv(
                        self.file_object,
                        sep=d,
                        nrows=limit,
                        on_bad_lines='skip',
                        low_memory=False
                    )
                    logger.info(f"✅ Read {len(df)} rows with delimiter '{d}'")
                    return self._normalize_columns(df)
                except Exception as e:
                    logger.warning(f"Failed with delimiter '{d}': {str(e)}")
                    self.file_object.seek(0)
                    continue

        try:
            logger.info(f"📖 Reading with default CSV reader")
            df = pd.read_csv(
                self.file_object,
                nrows=limit,
                on_bad_lines='skip',
                low_memory=False
            )
            logger.info(f"✅ Read {len(df)} rows")
            return self._normalize_columns(df)
        except Exception:
            logger.warning(f"Trying legacy error_bad_lines")
            self.file_object.seek(0)
            df = pd.read_csv(
                self.file_object,
                nrows=limit,
                error_bad_lines=False,
                warn_bad_lines=True,
                low_memory=False
            )
            logger.info(f"✅ Read {len(df)} rows (legacy)")
            return self._normalize_columns(df)

    def read_dbf(self, limit):
        temp = f"/tmp/{uuid.uuid4().hex}_{self.filename}"
        try:
            with open(temp, "wb") as f:
                self.file_object.seek(0)
                f.write(self.file_object.read())

            table = dbf.Table(temp)
            table.open()
            rows = []
            for i, r in enumerate(table):
                if limit and i >= limit:
                    break
                rows.append(dict(r))
            table.close()
            df = pd.DataFrame(rows)
            return self._normalize_columns(df)
        finally:
            if os.path.exists(temp):
                os.remove(temp)

    def read_sqlite(self, table_name, limit):
        """
        ✅ FIXED: Now logs available tables and handles selection better
        """
        temp = f"/tmp/{uuid.uuid4().hex}_{self.filename}"
        try:
            with open(temp, "wb") as f:
                self.file_object.seek(0)
                f.write(self.file_object.read())

            conn = sqlite3.connect(temp)

            # Get all tables (excluding system tables)
            tables_result = pd.read_sql(
                "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';",
                conn
            )
            tables = tables_result['name'].tolist() if len(tables_result) > 0 else []

            logger.info(f"📊 SQLite file has {len(tables)} table(s): {tables}")

            if not tables:
                conn.close()
                raise Exception("No tables found in SQLite database")

            # Determine which table to read
            if table_name:
                if table_name not in tables:
                    conn.close()
                    raise Exception(
                        f"Table '{table_name}' not found. "
                        f"Available tables: {tables}"
                    )
                target_table = table_name
                logger.info(f"📖 Reading selected table: '{table_name}'")
            else:
                target_table = tables[0]
                logger.info(f"📖 Reading first table: '{target_table}'")

            q = f"SELECT * FROM {target_table} LIMIT {limit}" if limit else f"SELECT * FROM {target_table}"
            df = pd.read_sql(q, conn)
            conn.close()

            logger.info(f"✅ Successfully read {len(df)} rows from SQLite table")
            return self._normalize_columns(df)

        except Exception as e:
            logger.error(f"Error reading SQLite file: {str(e)}")
            raise
        finally:
            if os.path.exists(temp):
                os.remove(temp)

    def read_access(self, table_name, limit):
        """
        ✅ DOCKER COMPATIBLE: Uses mdb-tools on Linux, pyodbc on Windows
        Works on both platforms automatically
        """
        temp = f"/tmp/{uuid.uuid4().hex}_{self.filename}"
        try:
            with open(temp, "wb") as f:
                self.file_object.seek(0)
                f.write(self.file_object.read())

            # Try pyodbc first (Windows), fall back to mdb-tools (Linux/Docker)
            try:
                conn = pyodbc.connect(
                    r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
                    f'DBQ={temp};'
                )
                cursor = conn.cursor()
                tables = [t.table_name for t in cursor.tables(tableType='TABLE')]

                logger.info(f"📊 Access file has {len(tables)} table(s): {tables}")

                if not tables:
                    conn.close()
                    raise Exception("No tables found in Access database")

                if table_name:
                    if table_name not in tables:
                        conn.close()
                        raise Exception(
                            f"Table '{table_name}' not found. Available tables: {tables}"
                        )
                    target_table = table_name
                else:
                    target_table = tables[0]

                logger.info(f"📖 Reading table: '{target_table}' (using pyodbc)")

                q = f"SELECT TOP {limit} * FROM [{target_table}]" if limit else f"SELECT * FROM [{target_table}]"
                df = pd.read_sql(q, conn)
                conn.close()

            except Exception as pyodbc_error:
                # Fall back to mdb-tools (Linux/Docker)
                logger.info(f"⚠️ pyodbc not available, using mdb-tools: {str(pyodbc_error)}")

                # Get table list using mdb-tools
                result = subprocess.run(
                    ['mdb-tables', '-1', temp],
                    capture_output=True,
                    text=True,
                    check=True
                )
                tables = [t.strip() for t in result.stdout.strip().split('\n') if t.strip()]

                logger.info(f"📊 Access file has {len(tables)} table(s): {tables}")

                if not tables:
                    raise Exception("No tables found in Access database")

                if table_name:
                    if table_name not in tables:
                        raise Exception(
                            f"Table '{table_name}' not found. Available tables: {tables}"
                        )
                    target_table = table_name
                else:
                    target_table = tables[0]

                logger.info(f"📖 Reading table: '{target_table}' (using mdb-tools)")

                # Export table to CSV using mdb-export
                result = subprocess.run(
                    ['mdb-export', temp, target_table],
                    capture_output=True,
                    text=True,
                    check=True
                )

                # Parse CSV output
                csv_data = StringIO(result.stdout)
                df = pd.read_csv(csv_data)

                if limit:
                    df = df.head(limit)

            logger.info(f"✅ Successfully read {len(df)} rows from Access table")
            return self._normalize_columns(df)

        except subprocess.CalledProcessError as e:
            logger.error(f"mdb-tools error: {e.stderr}")
            raise Exception(f"Failed to read Access file with mdb-tools: {e.stderr}")
        except Exception as e:
            logger.error(f"Error reading Access file: {str(e)}")
            raise Exception(f"Access file read failed: {str(e)}")
        finally:
            if os.path.exists(temp):
                os.remove(temp)

    # ==================== HELPER METHODS ====================

    def get_excel_sheets(self):
        """
        ✅ NEW: Get list of all sheets in Excel file

        Returns:
            list or None: Sheet names, or None if not an Excel file
        """
        if self.ext not in ['.xls', '.xlsx']:
            return None

        try:
            self.file_object.seek(0)
            xl_file = pd.ExcelFile(self.file_object)
            sheets = xl_file.sheet_names
            xl_file.close()
            logger.info(f"📊 Found {len(sheets)} Excel sheets: {sheets}")
            return sheets
        except Exception as e:
            logger.error(f"Error getting Excel sheet names: {str(e)}")
            return None

    def get_sqlite_tables(self):
        """
        ✅ NEW: Get list of all tables in SQLite file

        Returns:
            list or None: Table names, or None if not a SQLite file
        """
        if self.ext not in ['.db', '.sqlite', '.sqlite3']:
            return None

        temp = f"/tmp/{uuid.uuid4().hex}_{self.filename}"
        try:
            with open(temp, "wb") as f:
                self.file_object.seek(0)
                f.write(self.file_object.read())

            conn = sqlite3.connect(temp)
            cursor = conn.cursor()
            tables = cursor.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';"
            ).fetchall()
            table_names = [t[0] for t in tables]
            conn.close()

            logger.info(f"📊 Found {len(table_names)} SQLite tables: {table_names}")
            return table_names
        except Exception as e:
            logger.error(f"Error getting SQLite table names: {str(e)}")
            return None
        finally:
            if os.path.exists(temp):
                os.remove(temp)

    def get_access_tables(self):
        """
        ✅ DOCKER COMPATIBLE: Get list of all tables in Access file
        Works on both Windows (pyodbc) and Linux (mdb-tools)
        """
        if self.ext not in ['.mdb', '.accdb']:
            return None

        temp = f"/tmp/{uuid.uuid4().hex}_{self.filename}"
        try:
            with open(temp, "wb") as f:
                self.file_object.seek(0)
                f.write(self.file_object.read())

            try:
                # Try Windows/pyodbc first
                conn = pyodbc.connect(
                    r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
                    f'DBQ={temp};'
                )
                cursor = conn.cursor()
                tables = [t.table_name for t in cursor.tables(tableType='TABLE')]
                conn.close()
                logger.info(f"📊 Found {len(tables)} Access tables (pyodbc): {tables}")
                return tables

            except Exception:
                # Fall back to mdb-tools (Linux/Docker)
                logger.info(f"⚠️ pyodbc not available, using mdb-tools")

                result = subprocess.run(
                    ['mdb-tables', '-1', temp],
                    capture_output=True,
                    text=True,
                    check=True
                )
                tables = [t.strip() for t in result.stdout.strip().split('\n') if t.strip()]
                logger.info(f"📊 Found {len(tables)} Access tables (mdb-tools): {tables}")
                return tables

        except subprocess.CalledProcessError as e:
            logger.error(f"mdb-tools error: {e.stderr}")
            return None
        except Exception as e:
            logger.error(f"Error getting Access table names: {str(e)}")
            return None
        finally:
            if os.path.exists(temp):
                os.remove(temp)

    def _normalize_columns(self, df):
        """Ensure consistent column naming and strip whitespace"""
        needs_normalization = False

        # First, strip whitespace from all column names
        df.columns = [str(col).strip() for col in df.columns]

        for col in df.columns:
            if isinstance(col, str) and ('unnamed' in col.lower() or col.strip() == ''):
                needs_normalization = True
                break
            if isinstance(col, (int, np.integer)):
                needs_normalization = True
                break

        if needs_normalization:
            new_columns = [f"Column_{i}" for i in range(len(df.columns))]
            logger.info(f"🔧 Normalized columns to {new_columns}")
            df.columns = new_columns

        return df

    def get_row_count(self, table_name=None, sheet_name=None):
        """
        ✅ ENHANCED: Fast row counting with table/sheet support
        """
        try:
            if self.ext == '.csv':
                return self._count_csv_rows()
            elif self.ext in ['.xls', '.xlsx']:
                return self._count_excel_rows(sheet_name)
            elif self.ext == '.txt':
                return self._count_txt_rows()
            elif self.ext == '.dbf':
                return self._count_dbf_rows()
            elif self.ext in ['.db', '.sqlite', '.sqlite3']:
                return self._count_sqlite_rows(table_name)
            elif self.ext in ['.mdb', '.accdb']:
                return self._count_access_rows(table_name)
            else:
                return None
        except Exception as e:
            logger.warning(f"Could not count rows: {str(e)}")
            return None

    def _count_csv_rows(self):
        self.file_object.seek(0)
        count = sum(1 for _ in self.file_object) - 1
        self.file_object.seek(0)
        return count

    def _count_excel_rows(self, sheet_name=None):
        """✅ ENHANCED: Count rows in specific sheet"""
        import openpyxl
        self.file_object.seek(0)
        wb = openpyxl.load_workbook(self.file_object, read_only=True, data_only=True)

        if sheet_name and sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.active

        count = sheet.max_row - 1 if sheet.max_row > 0 else 0
        wb.close()
        return count

    def _count_txt_rows(self):
        return self._count_csv_rows()

    def _count_dbf_rows(self):
        temp = f"/tmp/{uuid.uuid4().hex}_{self.filename}"
        try:
            with open(temp, "wb") as f:
                self.file_object.seek(0)
                f.write(self.file_object.read())

            table = dbf.Table(temp)
            table.open()
            count = len(table)
            table.close()
            return count
        finally:
            if os.path.exists(temp):
                os.remove(temp)

    def _count_sqlite_rows(self, table_name=None):
        """✅ ENHANCED: Count rows in specific table"""
        temp = f"/tmp/{uuid.uuid4().hex}_{self.filename}"
        try:
            with open(temp, "wb") as f:
                self.file_object.seek(0)
                f.write(self.file_object.read())

            conn = sqlite3.connect(temp)
            cursor = conn.cursor()

            if not table_name:
                tables = cursor.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';").fetchall()
                if not tables:
                    conn.close()
                    return 0
                table_name = tables[0][0]

            count = cursor.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
            conn.close()
            return count
        finally:
            if os.path.exists(temp):
                os.remove(temp)

    def _count_access_rows(self, table_name=None):
        """✅ DOCKER COMPATIBLE: Count rows in specific table"""
        temp = f"/tmp/{uuid.uuid4().hex}_{self.filename}"
        try:
            with open(temp, "wb") as f:
                self.file_object.seek(0)
                f.write(self.file_object.read())

            try:
                # Try Windows/pyodbc first
                conn = pyodbc.connect(
                    r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
                    f'DBQ={temp};'
                )
                cursor = conn.cursor()

                if not table_name:
                    tables = [t.table_name for t in cursor.tables(tableType='TABLE')]
                    if not tables:
                        conn.close()
                        return 0
                    table_name = tables[0]

                count = cursor.execute(f"SELECT COUNT(*) FROM [{table_name}]").fetchone()[0]
                conn.close()
                return count

            except Exception:
                # Fall back to mdb-tools
                if not table_name:
                    result = subprocess.run(
                        ['mdb-tables', '-1', temp],
                        capture_output=True,
                        text=True,
                        check=True
                    )
                    tables = [t.strip() for t in result.stdout.strip().split('\n') if t.strip()]
                    if not tables:
                        return 0
                    table_name = tables[0]

                # Export and count rows
                result = subprocess.run(
                    ['mdb-export', temp, table_name],
                    capture_output=True,
                    text=True,
                    check=True
                )
                # Count lines minus header
                count = len(result.stdout.strip().split('\n')) - 1
                return max(0, count)

        except Exception as e:
            logger.warning(f"Could not count Access rows: {str(e)}")
            return None
        finally:
            if os.path.exists(temp):
                os.remove(temp)


# =============== MAPPING HELPERS ============
def get_mongodb_mapping():
    doc = MAPPING_COLLECTION.find_one() or {}
    doc.pop("_id", None)
    return doc


def save_user_mappings_to_db(user_mapping):
    if not user_mapping:
        return

    try:
        existing_doc = MAPPING_COLLECTION.find_one() or {}
        existing_doc.pop("_id", None)

        for db_key, file_columns in user_mapping.items():
            if not isinstance(file_columns, list):
                file_columns = [file_columns]

            existing = existing_doc.get(db_key, [])
            if not isinstance(existing, list):
                existing = [existing]

            new_columns = [c for c in file_columns if c not in existing]

            if new_columns:
                MAPPING_COLLECTION.update_one(
                    {},
                    {"$addToSet": {db_key: {"$each": new_columns}}},
                    upsert=True
                )

                logger.info(f"📝 Added mapping → {db_key}: {new_columns}")

    except Exception as e:
        logger.error(f"❌ Error saving user mappings: {str(e)}")
        raise


def check_column_mapping(file_cols, db_map):
    mapped, unmapped = {}, []
    for fc in file_cols:
        found = False
        for target, sources in db_map.items():
            if fc in sources:
                mapped[fc] = target
                found = True
                break
        if not found:
            unmapped.append(fc)
    return mapped, unmapped, list(db_map.keys())


# =============== DATA CLEANING ==============
def merge_columns(df, mapping):
    """
    ✅ SMART MERGE: Automatically detects and merges ALL address columns
    For Address field: Auto-includes Permanent + Local + Alternative address columns
    """
    result_df = pd.DataFrame()

    logger.info("=" * 80)
    logger.info("🔧 MERGE_COLUMNS FUNCTION STARTED (SMART ADDRESS MODE)")
    logger.info(f"📊 Input DataFrame shape: {df.shape}")
    logger.info(f"📋 Input DataFrame columns (original): {df.columns.tolist()}")

    # Normalize column names by stripping whitespace
    df.columns = [str(col).strip() for col in df.columns]
    logger.info(f"📋 Input DataFrame columns (normalized): {df.columns.tolist()}")
    logger.info("=" * 80)

    for target_col, source_cols in mapping.items():
        if not isinstance(source_cols, list):
            source_cols = [source_cols]

        # Normalize source column names
        source_cols = [str(col).strip() for col in source_cols]

        logger.info(f"\n{'=' * 60}")
        logger.info(f"🎯 Processing target column: '{target_col}'")
        logger.info(f"📝 Source columns requested: {source_cols}")

        # ✅ SMART ADDRESS HANDLING - Auto-detect and merge ALL address columns
        if target_col == "Address":
            logger.info("🏠 SMART ADDRESS MODE ACTIVATED")

            # Define all possible address columns
            perm_cols = ['Permanent Add1', 'Permanent Add2', 'Permanent Add3',
                         'Permanent City', 'Permanent State', 'Permanent Postal Code']

            local_cols = ['Local Add1', 'Local Add2', 'Local Add3',
                          'Local City', 'Local State', 'Local Postal Code']

            alt_cols = ['House No_Flat No', 'Street Address_Name', 'Locality',
                        'City', 'State_UT', 'Postal Code']

            # Check which columns exist in the dataframe
            available_perm = [col for col in perm_cols if col in df.columns]
            available_local = [col for col in local_cols if col in df.columns]
            available_alt = [col for col in alt_cols if col in df.columns]

            # Check which have actual data
            def check_data_exists(cols):
                if not cols:
                    return False, 0
                count = 0
                for col in cols:
                    non_empty = (df[col].notna() & (df[col].astype(str).str.strip() != '')).sum()
                    count += non_empty
                return count > 0, count

            perm_has_data, perm_count = check_data_exists(available_perm)
            local_has_data, local_count = check_data_exists(available_local)
            alt_has_data, alt_count = check_data_exists(available_alt)

            logger.info(f"🔍 ADDRESS FIELD ANALYSIS:")
            logger.info(
                f"   Permanent Address: {len(available_perm)} cols, {perm_count} non-empty values → {'✅ HAS DATA' if perm_has_data else '❌ EMPTY'}")
            logger.info(
                f"   Local Address: {len(available_local)} cols, {local_count} non-empty values → {'✅ HAS DATA' if local_has_data else '❌ EMPTY'}")
            logger.info(
                f"   Alternative Address: {len(available_alt)} cols, {alt_count} non-empty values → {'✅ HAS DATA' if alt_has_data else '❌ EMPTY'}")

            # Strategy: Merge ALL available address columns that have data
            columns_to_use = []

            if perm_has_data:
                columns_to_use.extend(available_perm)
                logger.info(f"   ✅ Including {len(available_perm)} Permanent Address columns")

            if local_has_data:
                columns_to_use.extend(available_local)
                logger.info(f"   ✅ Including {len(available_local)} Local Address columns")

            # Only use alternative if neither Permanent nor Local have data
            if alt_has_data and not (perm_has_data or local_has_data):
                columns_to_use.extend(available_alt)
                logger.info(f"   ✅ Including {len(available_alt)} Alternative Address columns")

            if not columns_to_use:
                logger.warning(f"   ⚠️ NO address columns have data! Using requested columns anyway")
                columns_to_use = [col for col in source_cols if col in df.columns]
            else:
                logger.info(f"🎯 SMART SELECTION: Merging {len(columns_to_use)} address columns with data")

            # Override source_cols with smart selection
            if columns_to_use:
                source_cols = columns_to_use

        # Continue with normal merge logic
        if len(source_cols) == 1:
            # Single column mapping
            if source_cols[0] in df.columns:
                result_df[target_col] = df[source_cols[0]].fillna('')
                sample = result_df[target_col].head(2).tolist()
                logger.info(f"✅ Single column mapped successfully")
                logger.info(f"   Sample values: {sample}")
            else:
                logger.warning(f"⚠️ Column '{source_cols[0]}' NOT FOUND in DataFrame")
                result_df[target_col] = ''
        else:
            # Multiple columns - merge them
            existing_cols = [col for col in source_cols if col in df.columns]
            missing_cols = [col for col in source_cols if col not in df.columns]

            logger.info(f"🔍 Column Analysis:")
            logger.info(f"   ✅ FOUND ({len(existing_cols)}): {existing_cols}")
            if missing_cols:
                logger.info(f"   ❌ MISSING ({len(missing_cols)}): {missing_cols}")

            if not existing_cols:
                logger.warning(f"⚠️ CRITICAL: NO columns found for '{target_col}'!")
                result_df[target_col] = ''
            else:
                logger.info(f"📝 Proceeding to merge {len(existing_cols)} columns")

                # Show sample data
                for i, col in enumerate(existing_cols[:3]):
                    sample_vals = df[col].head(2).tolist()
                    logger.info(f"   📊 '{col}' samples: {sample_vals}")
                if len(existing_cols) > 3:
                    logger.info(f"   ... and {len(existing_cols) - 3} more columns")

                # Convert to string and handle NaN
                selected_df = df[existing_cols].fillna('').astype(str)

                logger.info(f"🔧 Selected DataFrame shape: {selected_df.shape}")

                # Check data availability
                non_empty_per_col = {}
                for col in existing_cols:
                    non_empty_count = (df[col].notna() & (df[col].astype(str).str.strip() != '')).sum()
                    non_empty_per_col[col] = non_empty_count

                logger.info(f"🔍 DIAGNOSTIC - Non-empty counts:")
                total_non_empty = 0
                for col, count in non_empty_per_col.items():
                    pct = (count / len(df) * 100) if len(df) > 0 else 0
                    total_non_empty += count
                    logger.info(f"   '{col}': {count}/{len(df)} ({pct:.1f}%)")

                if total_non_empty == 0:
                    logger.error(f"🚨 ALL columns are empty - this is a DATA ISSUE!")

                # Merge row-wise
                result_df[target_col] = selected_df.apply(
                    lambda row: ', '.join([
                        val.strip()
                        for val in row
                        if val.strip() != '' and val.strip().lower() not in ('nan', 'none', 'null')
                    ]),
                    axis=1
                )

                # Show results
                sample_result = result_df[target_col].head(3).tolist()
                logger.info(f"✅ MERGED samples: {sample_result}")

                empty_count = (result_df[target_col] == '').sum()
                total_count = len(result_df[target_col])
                non_empty_count = total_count - empty_count

                logger.info(
                    f"📊 Final Stats: {non_empty_count}/{total_count} rows have data ({100 * non_empty_count / total_count:.1f}%)")

                if empty_count == total_count:
                    logger.error(f"🚨 ERROR: ALL rows empty after merge!")

    logger.info("\n" + "=" * 80)
    logger.info("🎯 MERGE COMPLETED")
    logger.info(f"📊 Output shape: {result_df.shape}")
    logger.info("=" * 80 + "\n")

    return result_df


def convert_dates(df):
    """Convert dates"""
    for col in ['DOB', 'DOA']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce", dayfirst=True)
            df[col] = df[col].apply(lambda x: x.to_pydatetime() if pd.notna(x) else None)
    return df


def clean_for_mongo(record):
    """Clean record for MongoDB"""
    cleaned = {}
    for key, value in record.items():
        if pd.isna(value):
            cleaned[key] = None
        elif isinstance(value, (np.integer, np.int64, np.int32)):
            cleaned[key] = int(value)
        elif isinstance(value, (np.floating, np.float64, np.float32)):
            cleaned[key] = float(value)
        elif isinstance(value, np.bool_):
            cleaned[key] = bool(value)
        elif isinstance(value, pd.Timestamp):
            cleaned[key] = value.to_pydatetime()
        else:
            cleaned[key] = value
    return cleaned


def process_chunk_bulk(chunk_df, operation, chunk_num, total_chunks, import_id, seq_id):
    """
    ⚡ OPTIMIZED: Process chunk with progress tracking and seq_id addition
    ✅ SMART DUPLICATE HANDLING:
    - Same group+subgroup → SKIP (do nothing)
    - Different group/subgroup → ADD seq_id to existing record (update)
    - New record → INSERT with seq_id
    """
    col = WATCHLIST_COLLECTION
    ins, upd, skip = 0, 0, 0
    operations = []

    for r in chunk_df.to_dict('records'):
        r = clean_for_mongo(r)

        if not r.get("Number"):
            skip += 1
            continue

        r["_id"] = str(r["Number"])

        # ✅ Check if record exists and if it already has this seq_id
        existing_record = col.find_one({"_id": r["_id"]}, {"seq_id": 1})

        if existing_record:
            existing_seq_ids = existing_record.get("seq_id", [])

            if seq_id in existing_seq_ids:
                # ✅ SKIP: Same group+subgroup already has this record
                skip += 1
                continue
            else:
                # ✅ UPDATE: Different group/subgroup - add new seq_id
                operations.append({
                    "type": "update_seq_id_only",
                    "record_id": r["_id"],
                    "seq_id": seq_id
                })
                upd += 1
                continue

        # ✅ NEW RECORD: Insert with seq_id
        r["seq_id"] = [seq_id]

        if operation == "insert":
            operations.append({
                "type": "insert",
                "data": r
            })
        else:
            operations.append({
                "type": "upsert",
                "data": r
            })

    # Execute bulk operations
    try:
        # Separate operations by type
        insert_ops = [op["data"] for op in operations if op.get("type") == "insert"]
        upsert_ops = []
        seq_id_update_ops = []

        for op in operations:
            if op.get("type") == "upsert":
                upsert_ops.append(
                    UpdateOne(
                        {"_id": op["data"]["_id"]},
                        {"$set": op["data"]},
                        upsert=True
                    )
                )
            elif op.get("type") == "update_seq_id_only":
                seq_id_update_ops.append(
                    UpdateOne(
                        {"_id": op["record_id"]},
                        {"$addToSet": {"seq_id": op["seq_id"]}}
                    )
                )

        # Execute inserts
        if insert_ops:
            try:
                result = col.insert_many(insert_ops, ordered=False)
                ins = len(result.inserted_ids)
            except errors.BulkWriteError as bwe:
                ins = bwe.details.get('nInserted', 0)
                write_errors = bwe.details.get('writeErrors', [])
                duplicate_errors = sum(1 for err in write_errors if err.get('code') == 11000)
                skip += duplicate_errors

                if duplicate_errors > 0:
                    logger.info(f"Chunk {chunk_num}: {ins} inserted, {duplicate_errors} duplicates skipped")

        # Execute upserts
        if upsert_ops:
            result = col.bulk_write(upsert_ops, ordered=False)
            ins += result.upserted_count

        # Execute seq_id updates
        if seq_id_update_ops:
            result = col.bulk_write(seq_id_update_ops, ordered=False)
            actual_modified = result.modified_count

            if actual_modified != len(seq_id_update_ops):
                logger.warning(
                    f"⚠️ Count mismatch: Expected {len(seq_id_update_ops)} updates, "
                    f"but DB modified {actual_modified}"
                )

    except Exception as e:
        logger.error(f"Bulk operation error in chunk {chunk_num}: {str(e)}")

    # Update progress
    if import_id:
        with import_progress_lock:
            if import_id in import_progress:
                import_progress[import_id]["completed_chunks"] += 1
                import_progress[import_id]["inserted"] += ins
                import_progress[import_id]["updated"] += upd
                import_progress[import_id]["skipped"] += skip

                progress_pct = (import_progress[import_id]["completed_chunks"] / total_chunks) * 100
                import_progress[import_id]["progress_percent"] = round(progress_pct, 1)

    return {"inserted": ins, "updated": upd, "skipped": skip}


def run_import_in_background(df, operation, import_id, file_id, seq_id, group_name, sub_group_name, filename,
                             description):
    """
    ⚡ Run the actual import in a background thread
    ✅ ENHANCED: Prevent duplicate imports for same group+subgroup
    """
    try:
        logger.info(f"🚀 Starting background import (import_id={import_id})")
        logger.info(f"🔖 Group: {group_name}, Sub-Group: {sub_group_name}, seq_id: {seq_id}")
        logger.info(f"📝 Description: {description}")
        logger.info(f"📁 Filename: {filename}")

        # ✅ Check if this exact group+subgroup combination already imported ANY file
        existing_nexus = WATCHLIST_NEXUS_COLLECTION.find_one({
            "Group_Name": group_name,
            "Sub_Group_Name": sub_group_name
        })

        if existing_nexus:
            logger.warning(f"🚫 DUPLICATE IMPORT DETECTED!")
            logger.warning(f"   File: '{filename}'")
            logger.warning(f"   Group: {group_name} / {sub_group_name}")
            logger.warning(f"   Previous import: {existing_nexus.get('InsertedAt')}")
            logger.warning(f"   seq_id: {existing_nexus.get('_id')}")
            logger.warning(f"   Records inserted: {existing_nexus.get('Inserted', 0)}")

            # Mark import as duplicate/skipped
            with import_progress_lock:
                if import_id in import_progress:
                    import_progress[import_id]["status"] = "duplicate"
                    import_progress[import_id]["progress_percent"] = 100
                    import_progress[import_id]["message"] = (
                        f"⚠️ File '{filename}' not imported - group '{group_name}/{sub_group_name}' "
                        f"already has data from previous import on "
                        f"{existing_nexus.get('InsertedAt').strftime('%Y-%m-%d %H:%M:%S') if existing_nexus.get('InsertedAt') else 'unknown date'}"
                    )
                    import_progress[import_id]["existing_nexus"] = {
                        "seq_id": existing_nexus.get('_id'),
                        "inserted_at": existing_nexus.get('InsertedAt'),
                        "records_inserted": existing_nexus.get('Inserted', 0),
                        "group_name": group_name,
                        "sub_group_name": sub_group_name
                    }
                    import_progress[import_id]["final_stats"] = {
                        "inserted": 0,
                        "updated": 0,
                        "skipped": existing_nexus.get('Inserted', 0)
                    }

            # Clean up
            pickle_path = f"{TEMP_DIR}/{file_id}.pkl"
            if os.path.exists(pickle_path):
                os.remove(pickle_path)

            with background_save_lock:
                background_save_status.pop(file_id, None)

            logger.info(f"⏭️ Import skipped - duplicate detected")
            return

        # ✅ Create Nexus record (new import)
        try:
            nexus_record = {
                "_id": seq_id,
                "Group_Name": group_name,
                "Sub_Group_Name": sub_group_name,
                "FileName": filename,
                "Description": description or "",  # ✅ Use empty string if None
                "Inserted": 0,
                "Duplicate": 0,
                "Updated": 0,
                "InsertedAt": datetime.now(),
                "Year": datetime.now().year
            }

            WATCHLIST_NEXUS_COLLECTION.insert_one(nexus_record)
            logger.info(f"✅ Nexus record created with seq_id: {seq_id}")
            logger.info(f"   Description: '{description}'")
        except errors.DuplicateKeyError:
            logger.warning(f"⚠️ Nexus record with seq_id {seq_id} already exists")
        except Exception as e:
            logger.error(f"❌ Failed to create Nexus record: {str(e)}")
            raise

        # Import data with seq_id and smart duplicate handling
        stats = insert_to_mongo_fast(df, operation, import_id, seq_id)

        # ✅ Update Nexus record with final stats
        try:
            WATCHLIST_NEXUS_COLLECTION.update_one(
                {"_id": seq_id},
                {
                    "$set": {
                        "Inserted": stats["inserted"],
                        "Duplicate": stats["skipped"],
                        "Updated": stats["updated"]
                    }
                }
            )
            logger.info(f"✅ Updated Nexus record:")
            logger.info(f"   - Inserted: {stats['inserted']} new records")
            logger.info(f"   - Duplicate: {stats['skipped']} (same group+subgroup - SKIPPED)")
            logger.info(f"   - Updated: {stats['updated']} (added seq_id to different group/subgroup)")
        except Exception as e:
            logger.warning(f"⚠️ Failed to update Nexus record: {str(e)}")

        # Mark as complete
        with import_progress_lock:
            if import_id in import_progress:
                import_progress[import_id]["final_stats"] = stats
                import_progress[import_id]["status"] = "completed"
                import_progress[import_id]["seq_id"] = seq_id

        # Clean up pickle file
        pickle_path = f"{TEMP_DIR}/{file_id}.pkl"
        if os.path.exists(pickle_path):
            os.remove(pickle_path)
            logger.info(f"🗑️ Cleaned up pickle file")

        # Clean up status tracking
        with background_save_lock:
            background_save_status.pop(file_id, None)

        logger.info(f"✅ Background import completed: {stats}")

    except Exception as e:
        logger.error(f"❌ Background import failed: {str(e)}", exc_info=True)

        with import_progress_lock:
            if import_id in import_progress:
                import_progress[import_id]["status"] = "failed"
                import_progress[import_id]["error"] = str(e)


def insert_to_mongo_fast(df, op, import_id=None, seq_id=None):
    """
    ⚡ OPTIMIZED: Faster parallel insertion with seq_id
    """
    total_rows = len(df)
    logger.info(f"⚡ Starting fast import of {total_rows} rows using {MAX_WORKERS} threads")
    logger.info(f"🔖 Using seq_id: {seq_id}")

    chunks = [df[i:i + CHUNK_SIZE] for i in range(0, total_rows, CHUNK_SIZE)]
    total_chunks = len(chunks)
    logger.info(f"📦 Split into {total_chunks} chunks of ~{CHUNK_SIZE} rows each")

    if import_id:
        with import_progress_lock:
            if import_id in import_progress:
                import_progress[import_id].update({
                    "total_chunks": total_chunks,
                    "status": "processing"
                })
            else:
                logger.warning(f"⚠️ Progress for {import_id} not found, creating new")
                import_progress[import_id] = {
                    "total_rows": total_rows,
                    "total_chunks": total_chunks,
                    "completed_chunks": 0,
                    "inserted": 0,
                    "updated": 0,
                    "skipped": 0,
                    "progress_percent": 0,
                    "status": "processing"
                }

    total_stats = {"inserted": 0, "updated": 0, "skipped": 0}

    start_time = time.time()

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_chunk = {
            executor.submit(process_chunk_bulk, chunk, op, i, total_chunks, import_id, seq_id): i
            for i, chunk in enumerate(chunks)
        }

        for future in as_completed(future_to_chunk):
            chunk_num = future_to_chunk[future]
            try:
                stats = future.result()
                total_stats["inserted"] += stats["inserted"]
                total_stats["updated"] += stats["updated"]
                total_stats["skipped"] += stats["skipped"]

                # Log every 10 chunks
                if (chunk_num + 1) % 10 == 0:
                    logger.info(f"✅ Chunk {chunk_num + 1}/{total_chunks} completed")
            except Exception as e:
                logger.error(f"❌ Chunk {chunk_num + 1} failed: {str(e)}")
                total_stats["skipped"] += len(chunks[chunk_num])

    elapsed_time = time.time() - start_time
    rows_per_sec = total_rows / elapsed_time if elapsed_time > 0 else 0

    logger.info(f"🎉 Import completed: {total_stats}")
    logger.info(f"⏱️ Time: {elapsed_time:.2f}s ({rows_per_sec:.0f} rows/sec)")

    if import_id:
        with import_progress_lock:
            if import_id in import_progress:
                import_progress[import_id]["status"] = "completed"
                import_progress[import_id]["progress_percent"] = 100
                import_progress[import_id]["elapsed_time"] = round(elapsed_time, 2)
                import_progress[import_id]["rows_per_second"] = round(rows_per_sec, 0)

    return total_stats


def save_file_in_background(bio, filename, file_id):
    """
    ✅ ENHANCED: Automatically reads and combines ALL sheets/tables from multi-table files
    ✅ FIXED: Preserves description throughout the background save process
    """
    try:
        # ✅ Get existing status to preserve description
        with background_save_lock:
            existing_status = background_save_status.get(file_id, {})
            existing_description = existing_status.get("description", "")

            background_save_status[file_id] = {
                "status": "in_progress",
                "error": None,
                "filename": filename,
                "description": existing_description  # ✅ Preserve description
            }

        logger.info(f"🔄 Background: Reading file {filename} (file_id={file_id})")
        logger.info(f"📝 Preserving description: '{existing_description}'")

        reader = FileColumnReader(bio, filename)

        # Check if file has multiple tables/sheets
        multi_table_info = reader.get_all_tables_or_sheets()

        if multi_table_info and multi_table_info['count'] > 1:
            logger.info(
                f"📊 Multi-table file detected: {multi_table_info['count']} {multi_table_info['type']} tables/sheets")
            logger.info(f"   Tables/Sheets: {multi_table_info['tables']}")
            logger.info(f"🔗 Automatically combining all tables/sheets...")

            # Read and combine all tables/sheets
            all_dfs = []
            ext = Path(filename).suffix.lower()

            for i, table_or_sheet in enumerate(multi_table_info['tables']):
                try:
                    logger.info(f"   📖 Reading {i + 1}/{multi_table_info['count']}: '{table_or_sheet}'...")

                    # Reset bio for each read
                    bio.seek(0)
                    temp_reader = FileColumnReader(bio, filename)

                    if ext in ['.xls', '.xlsx']:
                        df = temp_reader.read_data(limit=None, sheet_name=table_or_sheet)
                    else:
                        df = temp_reader.read_data(limit=None, table_name=table_or_sheet)

                    # Add source column to track which table/sheet each row came from
                    df['_source'] = table_or_sheet

                    all_dfs.append(df)
                    logger.info(f"      ✅ Read {len(df)} rows, {len(df.columns) - 1} columns")

                except Exception as e:
                    logger.warning(f"      ⚠️ Failed to read '{table_or_sheet}': {str(e)}")
                    continue

            if not all_dfs:
                raise Exception("Failed to read any tables/sheets from the file")

            # Combine all DataFrames
            logger.info(f"🔗 Combining {len(all_dfs)} tables/sheets...")
            df_full = pd.concat(all_dfs, ignore_index=True, sort=False)

            logger.info(f"✅ Combined successfully!")
            logger.info(f"   Total rows: {len(df_full)}")
            logger.info(f"   Total unique columns: {len(df_full.columns)}")
            logger.info(f"   Sources: {df_full['_source'].unique().tolist()}")

        else:
            # Single table/sheet file - read normally
            logger.info(f"📄 Single-table file")
            df_full = reader.read_data(limit=None)

        logger.info(f"📖 Background: Read {len(df_full)} rows total")
        logger.info(f"📋 Background: DataFrame columns: {df_full.columns.tolist()}")

        pickle_path = f"{TEMP_DIR}/{file_id}.pkl"
        df_full.to_pickle(pickle_path)

        if not os.path.exists(pickle_path):
            raise Exception(f"Pickle file not created at {pickle_path}")

        file_size = os.path.getsize(pickle_path)
        logger.info(f"💾 Background: Saved to {pickle_path} ({file_size} bytes)")

        # ✅ Update status while preserving description
        with background_save_lock:
            background_save_status[file_id] = {
                "status": "completed",
                "rows": len(df_full),
                "columns": df_full.columns.tolist(),
                "file_size": file_size,
                "filename": filename,
                "multi_table_info": multi_table_info,
                "error": None,
                "description": existing_description  # ✅ Keep description here
            }

        logger.info(f"✅ Background save completed with description: '{existing_description}'")

    except Exception as e:
        error_msg = f"{str(e)}\n{traceback.format_exc()}"
        logger.error(f"❌ Background save failed (file_id={file_id}): {error_msg}")

        # ✅ Preserve description even on failure
        with background_save_lock:
            existing_status = background_save_status.get(file_id, {})
            existing_description = existing_status.get("description", "")

            background_save_status[file_id] = {
                "status": "failed",
                "error": error_msg,
                "filename": filename,
                "description": existing_description  # ✅ Keep description on error too
            }


# ================= MAIN API =================
class WatchlistFileUploadView(APIView):
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request):
        file_id = request.POST.get("file_id") or request.data.get("file_id")
        column_mapping = request.POST.get("column_mapping") or request.data.get("column_mapping")
        operation = request.POST.get("operation") or request.data.get("operation", "upsert")

        # ✅ Get Group Name, Sub Group Name, and Description
        group_name = request.POST.get("group_name") or request.data.get("group_name")
        sub_group_name = request.POST.get("sub_group_name") or request.data.get("sub_group_name")
        description = request.POST.get("description") or request.data.get("description", "")

        logger.info(f"Received request - file_id: {file_id}, operation: {operation}")
        logger.info(f"Initial description value: '{description}'")

        # ✅ If importing (file_id present) and description is empty, try to retrieve from cache
        if file_id and not description:
            with background_save_lock:
                save_status = background_save_status.get(file_id, {})
                cached_description = save_status.get("description", "")
                if cached_description:
                    description = cached_description
                    logger.info(f"📝 Retrieved description from cache: '{description}'")

        logger.info(f"📥 Final parameters:")
        logger.info(f"   - group_name: {group_name}")
        logger.info(f"   - sub_group_name: {sub_group_name}")
        logger.info(f"   - description: '{description}'")
        logger.info(f"   - operation: {operation}")

        # ============================================================
        # ✅ STEP 2 — START ASYNC IMPORT
        # ============================================================
        if file_id:
            # ✅ Validate Group Name and Sub Group Name
            if not group_name or not sub_group_name:
                return Response({
                    "error": "Group Name and Sub Group Name are required",
                    "missing": {
                        "group_name": not group_name,
                        "sub_group_name": not sub_group_name
                    }
                }, status=400)

            path = f"{TEMP_DIR}/{file_id}.pkl"

            # ⚡ WAIT for background save to complete (with timeout)
            max_wait_time = 60
            wait_interval = 0.5
            total_waited = 0

            while total_waited < max_wait_time:
                with background_save_lock:
                    save_status = background_save_status.get(file_id)

                if save_status:
                    if save_status["status"] == "completed":
                        break
                    elif save_status["status"] == "failed":
                        return Response({
                            "error": "Background file save failed",
                            "detail": save_status.get("error", "Unknown error"),
                            "file_id": file_id
                        }, status=500)
                    elif save_status["status"] == "in_progress":
                        time.sleep(wait_interval)
                        total_waited += wait_interval
                        continue
                else:
                    if os.path.exists(path):
                        break
                    else:
                        time.sleep(wait_interval)
                        total_waited += wait_interval

            # After waiting, check one more time
            with background_save_lock:
                save_status = background_save_status.get(file_id)

            if save_status and save_status["status"] == "in_progress":
                return Response({
                    "error": "File is still being processed (timeout)",
                    "message": "Large file is still being prepared. Please try again in a moment.",
                    "file_id": file_id,
                    "waited_seconds": total_waited
                }, status=202)

            # Check file exists
            if not os.path.exists(path):
                logger.error(f"❌ Pickle file not found: {path}")
                return Response({
                    "error": "Invalid or expired file_id",
                    "file_id": file_id
                }, status=400)

            if not column_mapping:
                return Response({"error": "column_mapping required"}, status=400)

            try:
                # Load dataset
                df = pd.read_pickle(path)
                logger.info(f"📊 Loaded {len(df)} rows")
                logger.info(f"📋 Loaded DataFrame columns: {df.columns.tolist()}")

                # Get filename from background save status
                filename = save_status.get("filename", "unknown_file")

                # Parse user mapping
                if isinstance(column_mapping, str):
                    user_mapping = json.loads(column_mapping)
                else:
                    user_mapping = column_mapping

                # Parse auto mapping
                auto_mapping = request.POST.get("auto_mapping") or request.data.get("auto_mapping")
                if isinstance(auto_mapping, str):
                    auto_mapping = json.loads(auto_mapping)
                elif not auto_mapping:
                    auto_mapping = {}

                logger.info(f"📋 User mapping: {user_mapping}")
                logger.info(f"🤖 Auto mapping: {auto_mapping}")

                # ✅ Combine mappings correctly
                final_mapping = {}

                # Process auto_mapping
                for db_key, file_cols in auto_mapping.items():
                    if not isinstance(file_cols, list):
                        file_cols = [file_cols]

                    if db_key not in final_mapping:
                        final_mapping[db_key] = []

                    for col in file_cols:
                        if col not in final_mapping[db_key]:
                            final_mapping[db_key].append(col)

                # Process user_mapping
                for db_key, file_cols in user_mapping.items():
                    if not isinstance(file_cols, list):
                        file_cols = [file_cols]

                    if db_key not in final_mapping:
                        final_mapping[db_key] = []

                    for col in file_cols:
                        if col not in final_mapping[db_key]:
                            final_mapping[db_key].append(col)

                logger.info(f"✅ Final mapping: {json.dumps(final_mapping, indent=2)}")

                # ✅ Save user mappings to MongoDB for future auto-mapping
                try:
                    save_user_mappings_to_db(user_mapping)
                    logger.info("✅ User mappings saved to database for future auto-mapping")
                except Exception as e:
                    logger.warning(f"⚠️ Failed to save user mappings to database: {str(e)}")

                # Validate required mapping
                if "Number" not in final_mapping or len(final_mapping["Number"]) == 0:
                    return Response({
                        "error": "Number column must be mapped",
                        "current_mapping": final_mapping
                    }, status=400)

                # Transform data
                logger.info("\n" + "=" * 80)
                logger.info("🚀 STARTING DATA TRANSFORMATION")
                logger.info("=" * 80)

                df = merge_columns(df, final_mapping)

                # Show first record after merge
                logger.info("\n" + "=" * 80)
                logger.info("🔍 FIRST RECORD AFTER MERGE:")
                if len(df) > 0:
                    first_record = df.iloc[0].to_dict()
                    for key, val in first_record.items():
                        val_preview = str(val)[:100] if val else ""
                        logger.info(f"   {key}: '{val_preview}'")
                logger.info("=" * 80 + "\n")

                df = convert_dates(df)

                # ⚡ Generate import_id and seq_id
                import_id = str(uuid.uuid4())
                seq_id = generate_hash_id(group_name, sub_group_name)

                logger.info(f"🔖 Generated seq_id: {seq_id} for {group_name}/{sub_group_name}")

                # ✅ Initialize progress BEFORE starting background thread
                with import_progress_lock:
                    import_progress[import_id] = {
                        "total_rows": len(df),
                        "total_chunks": 0,
                        "completed_chunks": 0,
                        "inserted": 0,
                        "updated": 0,
                        "skipped": 0,
                        "progress_percent": 0,
                        "status": "starting",
                        "group_name": group_name,
                        "sub_group_name": sub_group_name,
                        "seq_id": seq_id,
                        "description": description
                    }

                logger.info(f"✅ Progress tracking initialized for import_id: {import_id}")

                # ⚡ Start import in background
                executor = ThreadPoolExecutor(max_workers=1)
                executor.submit(
                    run_import_in_background,
                    df,
                    operation,
                    import_id,
                    file_id,
                    seq_id,
                    group_name,
                    sub_group_name,
                    filename,
                    description
                )
                executor.shutdown(wait=False)

                logger.info(f"✅ Import started in background (import_id={import_id})")

                # ⚡ Return immediately with import_id
                return Response({
                    "status": "import_started",
                    "message": "Import started in background",
                    "import_id": import_id,
                    "seq_id": seq_id,
                    "group_name": group_name,
                    "sub_group_name": sub_group_name,
                    "description": description,
                    "total_rows": len(df),
                    "estimated_time_seconds": round(len(df) / 10000, 1)
                })

            except Exception as e:
                logger.error(f"Import error: {str(e)}", exc_info=True)
                return Response({
                    "error": "Import failed",
                    "detail": str(e)
                }, status=500)

        # ============================================================
        # ⚡ STEP 1 — PREVIEW (✅ AUTO-COMBINES MULTI-TABLE FILES)
        # ============================================================
        uploaded_files = request.FILES.getlist("file")
        if not uploaded_files:
            return Response({"error": "No file uploaded"}, status=400)

        file = uploaded_files[0]
        ext = file.name.split('.')[-1].lower()

        if ext not in ALLOWED_EXTENSIONS:
            return Response({
                "error": "Unsupported file type",
                "allowed": list(ALLOWED_EXTENSIONS)
            }, status=400)

        try:
            start_time = time.time()

            file_content = file.read()
            bio = BytesIO(file_content)

            reader = FileColumnReader(bio, file.name)

            # ✅ Check if file has multiple tables/sheets
            multi_table_info = reader.get_all_tables_or_sheets()

            if multi_table_info and multi_table_info['count'] > 1:
                logger.info(f"📊 Multi-table file detected: {multi_table_info}")
                logger.info(f"🔗 Reading preview from ALL {multi_table_info['count']} tables/sheets...")

                # Read preview from ALL tables/sheets and combine
                all_previews = []
                for table_or_sheet in multi_table_info['tables']:
                    try:
                        bio.seek(0)
                        temp_reader = FileColumnReader(bio, file.name)

                        if ext in ['.xls', '.xlsx']:
                            df_temp = temp_reader.read_data(limit=PREVIEW_ROWS, sheet_name=table_or_sheet)
                        else:
                            df_temp = temp_reader.read_data(limit=PREVIEW_ROWS, table_name=table_or_sheet)

                        df_temp['_source'] = table_or_sheet
                        all_previews.append(df_temp)
                        logger.info(f"   ✅ Preview from '{table_or_sheet}': {len(df_temp)} rows")
                    except Exception as e:
                        logger.warning(f"   ⚠️ Failed to preview '{table_or_sheet}': {str(e)}")

                if all_previews:
                    df_preview = pd.concat(all_previews, ignore_index=True, sort=False)
                    logger.info(f"✅ Combined preview: {len(df_preview)} rows from {len(all_previews)} tables/sheets")
                else:
                    raise Exception("Failed to read preview from any table/sheet")
            else:
                # Single table/sheet - read normally
                logger.info(f"📖 Reading preview ({PREVIEW_ROWS} rows): {file.name}")
                df_preview = reader.read_data(limit=PREVIEW_ROWS)

            file_cols = df_preview.columns.tolist()
            logger.info(f"📋 Preview columns: {file_cols}")

            # Count total rows (this will be done in background for multi-table files)
            if multi_table_info and multi_table_info['count'] > 1:
                total_rows = "calculating..."
            else:
                bio_for_count = BytesIO(file_content)
                count_reader = FileColumnReader(bio_for_count, file.name)
                total_rows = count_reader.get_row_count()

            db_map = get_mongodb_mapping()
            mapped, unmapped, db_keys = check_column_mapping(file_cols, db_map)

            file_id = str(uuid.uuid4())

            # ✅ Store description for later import call
            logger.info(f"📝 Storing description for file_id {file_id}: '{description}'")
            with background_save_lock:
                background_save_status[file_id] = {
                    "status": "pending",
                    "description": description,  # ✅ Save description here
                    "filename": file.name
                }

            auto_mapped_dict = {}
            for file_col, db_key in mapped.items():
                if db_key not in auto_mapped_dict:
                    auto_mapped_dict[db_key] = []
                auto_mapped_dict[db_key].append(file_col)

            # ✅ SYNCHRONOUS full file save — eliminates the race condition where
            # the import request arrives before the background thread finishes.
            # The user waits here once (during preview) instead of getting stuck
            # when they click Import. Django's single thread is NOT blocked between
            # requests — it's blocked inside this one preview request which is fine.
            logger.info(f"💾 Reading & saving full file synchronously (file_id={file_id})...")
            save_start = time.time()
            try:
                bio_for_save = BytesIO(file_content)
                save_file_in_background(bio_for_save, file.name, file_id)
                save_elapsed = time.time() - save_start
                logger.info(f"✅ Full file saved in {save_elapsed:.2f}s — import is ready")
            except Exception as save_err:
                logger.error(f"❌ Synchronous file save failed: {save_err}", exc_info=True)
                # Mark as failed so import returns a clear error
                with background_save_lock:
                    background_save_status[file_id] = {
                        "status": "failed",
                        "error": str(save_err),
                        "filename": file.name,
                        "description": description
                    }

            elapsed_time = time.time() - start_time
            logger.info(f"⚡ Preview + save completed in {elapsed_time:.2f}s (file_id={file_id})")

            # ✅ Build response
            response_data = {
                "status": "preview",
                "file_ready": True,  # pickle is already saved — import can start immediately
                "file_id": file_id,
                "file_columns": file_cols,
                "auto_mapped": auto_mapped_dict,
                "unmapped_columns": unmapped,
                "database_keys": db_keys,
                "total_rows": total_rows,
                "sample_data": df_preview.fillna("").to_dict(orient="records"),
                "preview_time_seconds": round(elapsed_time, 2),
                "requires_group_info": True,
                "message": "⚠️ Please provide Group Name and Sub Group Name before importing"
            }

            # ✅ Add multi-table information if detected
            if multi_table_info and multi_table_info['count'] > 1:
                response_data["multi_table_detected"] = True
                response_data["table_info"] = multi_table_info

                if multi_table_info['type'] == 'excel':
                    response_data["message"] = (
                        f"✅ Excel file with {multi_table_info['count']} sheets detected. "
                        f"ALL sheets will be automatically combined and imported: "
                        f"{', '.join(multi_table_info['tables'])}"
                    )
                elif multi_table_info['type'] in ['sqlite', 'access']:
                    response_data["message"] = (
                        f"✅ Database with {multi_table_info['count']} tables detected. "
                        f"ALL tables will be automatically combined and imported: "
                        f"{', '.join(multi_table_info['tables'])}"
                    )

                logger.info(f"✅ Multi-table file will be auto-combined: {multi_table_info}")
            else:
                response_data["multi_table_detected"] = False

            return Response(response_data)

        except Exception as e:
            logger.error(f"Preview error: {str(e)}", exc_info=True)
            return Response({
                "error": "Failed to read file",
                "detail": str(e)
            }, status=500)


class ImportProgressViewWatchlist(APIView):
    """
    ⚡ Check import progress
    GET /import-progress/{import_id}/
    """

    def get(self, request, import_id):
        with import_progress_lock:
            doc = import_progress.get(import_id)

        if not doc:
            logger.warning(f"⚠️ Import ID not found: {import_id}")
            return Response({
                "error": "Import ID not found",
                "import_id": import_id,
                "message": "Import may not have started yet or import_id is invalid"
            }, status=404)

        return Response(doc)