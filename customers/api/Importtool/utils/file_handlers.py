import os
import io
import json
import shutil
import zipfile
import tempfile
from io import StringIO, BytesIO

import polars as pl
from abc import ABC, abstractmethod
from openpyxl import load_workbook
import csv
import chardet
import logging

logger = logging.getLogger(__name__)

# ============================================================
# Archive Support
# ============================================================

# Try to import RAR support
try:
    import rarfile as _rarfile_module
    _RAR_AVAILABLE = True
except ImportError:
    _rarfile_module = None
    _RAR_AVAILABLE = False

# Supported archive and CDR file extensions
_ARCHIVE_EXTS = {'.zip', '.rar'}
_CDR_EXTS     = {'.xlsx', '.xls', '.csv', '.txt', '.parquet'}


def get_rarfile():
    """Lazy-load rarfile with UnRAR tool detection."""
    global _RAR_AVAILABLE, _rarfile_module
    if not _RAR_AVAILABLE:
        try:
            import rarfile as _rf
            _rarfile_module = _rf
            _RAR_AVAILABLE = True
        except ImportError:
            return None

    try:
        import unrar.cffi
        _rarfile_module.UNRAR_LIB_PATH = unrar.cffi.lib_path()
        return _rarfile_module
    except Exception:
        pass

    candidates = (
        'unrar', 'UnRAR',
        r'C:\Program Files\WinRAR\UnRAR.exe',
        r'C:\Program Files (x86)\WinRAR\UnRAR.exe',
    )
    for candidate in candidates:
        path = shutil.which(candidate) or (candidate if os.path.isfile(candidate) else None)
        if path:
            _rarfile_module.UNRAR_TOOL = path
            return _rarfile_module

    return _rarfile_module


def is_archive(filename: str) -> bool:
    """Return True if filename has a ZIP or RAR extension."""
    return any(filename.lower().endswith(ext) for ext in _ARCHIVE_EXTS)


def is_cdr_file(filename: str) -> bool:
    """
    Return True if filename is a supported CDR data format.
    Handles double-extensions like .csv.xls and logs unrecognised extensions.
    """
    name_lower = filename.lower()

    # Double-extension check (.csv.xls, .csv.xlsx — CSV with misleading extension)
    for double_ext in ('.csv.xls', '.csv.xlsx'):
        if name_lower.endswith(double_ext):
            return True

    ext = os.path.splitext(name_lower)[1]
    matched = ext in _CDR_EXTS
    if not matched:
        logger.warning(
            f"[is_cdr_file] '{filename}' — extension '{ext}' NOT in {_CDR_EXTS}. "
            f"File will be skipped. Add the extension to _CDR_EXTS if it should be supported."
        )
    return matched


def extract_files_from_archive(archive_bytes, archive_name: str, _depth: int = 0):
    """
    Recursively extract CDR files from a ZIP or RAR archive.
    Logs every entry it encounters so skipped files are visible in the logs.
    Yields (filename, file_bytes) for every matched CDR file.
    """
    MAX_DEPTH = 10
    indent = '   ' * (_depth + 1)

    if _depth > MAX_DEPTH:
        logger.warning(f"{indent}Max archive nesting depth reached — skipping '{archive_name}'")
        return

    if isinstance(archive_bytes, bytes):
        file_obj = io.BytesIO(archive_bytes)
    else:
        file_obj = archive_bytes
        file_obj.seek(0)

    archive_lower = archive_name.lower()

    try:
        # ── RAR ──────────────────────────────────────────────
        if archive_lower.endswith('.rar'):
            rf_mod = get_rarfile()
            if rf_mod is None:
                logger.error(f"{indent}RAR support not available — skipping '{archive_name}'")
                return

            logger.info(f"{indent}Opening RAR: '{archive_name}'")

            # rarfile cannot reliably open a BytesIO object — it requires a real
            # file path. Write the archive bytes to a temp file first, then open
            # it by path, and always clean up afterwards.
            tmp_suffix = os.path.splitext(archive_name)[1] or '.rar'
            with tempfile.NamedTemporaryFile(suffix=tmp_suffix, delete=False) as tmp:
                tmp_path = tmp.name
                file_obj.seek(0)
                tmp.write(file_obj.read())

            try:
                with rf_mod.RarFile(tmp_path) as rf:
                    all_entries = rf.infolist()
                    logger.info(f"{indent}RAR has {len(all_entries)} entries:")

                    for entry in all_entries:
                        raw_entry_name = entry.filename
                        # RAR archives created on Windows use backslashes
                        normalised = raw_entry_name.replace('\\\\', '/')
                        base_name  = os.path.basename(normalised)

                        logger.info(
                            f"{indent}  entry='{raw_entry_name}' "
                            f"base='{base_name}' "
                            f"is_dir={entry.is_dir()} "
                            f"size={getattr(entry, 'file_size', '?')}")

                        if entry.is_dir():
                            continue

                        if not base_name:
                            logger.warning(f"{indent}    → empty basename, skipping")
                            continue

                        try:
                            raw_bytes = rf.read(raw_entry_name)
                        except Exception as e:
                            logger.error(f"{indent}    → could not read entry '{raw_entry_name}': {e}")
                            continue

                        if is_archive(base_name):
                            logger.info(f"{indent}    → nested archive, recursing into '{base_name}'")
                            yield from extract_files_from_archive(raw_bytes, base_name, _depth + 1)
                        elif is_cdr_file(base_name):
                            logger.info(f"{indent}    → ✅ yielding CDR file: '{base_name}'")
                            yield base_name, raw_bytes
                        else:
                            logger.warning(f"{indent}    → ⚠️  SKIPPED (unrecognised): '{base_name}'")
            finally:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
        # ── ZIP ──────────────────────────────────────────────
        elif archive_lower.endswith('.zip'):
            logger.info(f"{indent}Opening ZIP: '{archive_name}'")

            with zipfile.ZipFile(file_obj, 'r') as zf:
                all_entries = zf.infolist()
                logger.info(f"{indent}ZIP has {len(all_entries)} entries:")

                for entry in all_entries:
                    raw_entry_name = entry.filename
                    base_name      = os.path.basename(raw_entry_name.replace('\\', '/'))

                    logger.info(
                        f"{indent}  entry='{raw_entry_name}' "
                        f"base='{base_name}' "
                        f"is_dir={entry.is_dir()} "
                        f"size={entry.file_size}"
                    )

                    if entry.is_dir():
                        continue

                    if not base_name:
                        logger.warning(f"{indent}    → empty basename, skipping")
                        continue

                    try:
                        raw_bytes = zf.read(raw_entry_name)
                    except Exception as e:
                        logger.error(f"{indent}    → could not read entry '{raw_entry_name}': {e}")
                        continue

                    if is_archive(base_name):
                        logger.info(f"{indent}    → nested archive, recursing into '{base_name}'")
                        yield from extract_files_from_archive(raw_bytes, base_name, _depth + 1)
                    elif is_cdr_file(base_name):
                        logger.info(f"{indent}    → ✅ yielding CDR file: '{base_name}'")
                        yield base_name, raw_bytes
                    else:
                        logger.warning(f"{indent}    → ⚠️  SKIPPED (unrecognised): '{base_name}'")

        else:
            logger.error(f"{indent}Unknown archive format: '{archive_name}'")

    except zipfile.BadZipFile as e:
        logger.error(f"{indent}Bad ZIP file '{archive_name}': {e}")
    except Exception as e:
        logger.error(f"{indent}Could not open archive '{archive_name}': {e}", exc_info=True)


def expand_archive_files(file_data: list) -> list:
    """
    Expand any ZIP/RAR entries in file_data into their CDR file contents.
    Non-archive files are passed through unchanged.

    Args:
        file_data: list of (file_obj, filename) tuples

    Returns:
        list of (file_bytes: bytes, filename: str) ready for processing
    """
    expanded = []

    for file_obj, filename in file_data:
        if hasattr(file_obj, 'read'):
            file_obj.seek(0)
            file_bytes = file_obj.read()
        else:
            file_bytes = file_obj

        if is_archive(filename):
            logger.info(f"\n[ARCHIVE] Detected: '{filename}' — extracting CDR files...")
            extracted_count = 0
            try:
                for inner_name, inner_bytes in extract_files_from_archive(file_bytes, filename):
                    expanded.append((inner_bytes, inner_name))
                    extracted_count += 1

                logger.info(f"[ARCHIVE] ✅ {extracted_count} CDR file(s) extracted from '{filename}'")
                if extracted_count == 0:
                    logger.warning(
                        f"[ARCHIVE] ⚠️  No CDR files matched inside '{filename}'. "
                        f"Check logs above for the actual filenames/extensions found — "
                        f"add missing extensions to _CDR_EXTS if needed."
                    )
            except Exception as e:
                logger.error(f"[ARCHIVE] ❌ Failed to extract '{filename}': {e}", exc_info=True)
        else:
            expanded.append((file_bytes, filename))

    return expanded


# ============================================================
# Base Strategy Pattern
# ============================================================

class FileStrategy(ABC):
    @abstractmethod
    def read(self, file_source):
        pass


class FileHandler:
    def __init__(self, strategy: FileStrategy):
        self._strategy = strategy

    def set_strategy(self, strategy: FileStrategy):
        self._strategy = strategy

    def read_file(self, file_source):
        return self._strategy.read(file_source)


# ============================================================
# Memory-based Handlers
# ============================================================

class CSVMemoryHandler(FileStrategy):

    def _detect_encoding(self, file_obj):
        """Detect file encoding from a BytesIO object."""
        try:
            file_obj.seek(0)
            rawdata = file_obj.read(100000)
            file_obj.seek(0)
            return chardet.detect(rawdata).get('encoding', 'utf-8')
        except Exception:
            return 'utf-8'

    def read(self, file_obj):
        """
        Read CSV from a BytesIO / file-like object with full string preservation.
        Also accepts raw bytes (from archive extraction).
        """
        if isinstance(file_obj, (bytes, bytearray)):
            file_obj = io.BytesIO(file_obj)

        encoding = self._detect_encoding(file_obj)

        try:
            file_obj.seek(0)
            text_content = file_obj.read().decode(encoding, errors='ignore')

            rows = list(csv.reader(StringIO(text_content)))
            if not rows:
                raise ValueError("CSV file has no data")

            max_cols = max(len(r) for r in rows)
            data_dict = {f"col_{i}": [] for i in range(max_cols)}
            for row in rows:
                for i in range(max_cols):
                    data_dict[f"col_{i}"].append(str(row[i]) if i < len(row) else "")

            return pl.DataFrame(
                data_dict,
                schema_overrides={col: pl.Utf8 for col in data_dict}
            )

        except Exception as e:
            raise ValueError(f"Error reading CSV from memory: {e}") from e


class ExcelMemoryHandler(FileStrategy):

    def read(self, file_obj):
        """
        Read Excel from a BytesIO object.
        Handles merged cells properly (data_only=True, not read_only).
        Also accepts raw bytes (from archive extraction).
        """
        if isinstance(file_obj, (bytes, bytearray)):
            file_obj = io.BytesIO(file_obj)

        try:
            file_obj.seek(0)
            wb = load_workbook(file_obj, data_only=True)
            sheets = {}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                data = [tuple(cell.value for cell in row) for row in ws.iter_rows()]
                if not data:
                    continue

                max_cols = max(len(row) for row in data)
                if max_cols == 0:
                    continue

                normalized_data = [
                    (list(row) + [None] * (max_cols - len(row)))[:max_cols]
                    for row in data
                ]

                header = normalized_data[0]
                rows   = normalized_data[1:]

                header_clean = []
                seen_headers = {}
                for i, h in enumerate(header):
                    col_name = str(h).strip() if h is not None and str(h).strip() else f"col_{i}"
                    if col_name in seen_headers:
                        seen_headers[col_name] += 1
                        col_name = f"{col_name}_{seen_headers[col_name]}"
                    else:
                        seen_headers[col_name] = 0
                    header_clean.append(col_name)

                schema = {h: pl.Utf8 for h in header_clean}

                rows_clean = [
                    [
                        str(row[i]) if i < len(row) and row[i] is not None else ""
                        for i in range(max_cols)
                    ]
                    for row in rows
                ]

                if rows_clean:
                    sheets[sheet_name] = pl.DataFrame(rows_clean, schema=schema, orient="row")

            return sheets

        except Exception as e:
            raise ValueError(f"Cannot read Excel from memory: {e}") from e


# ============================================================
# Archive Memory Handler
# ============================================================

class ArchiveMemoryHandler(FileStrategy):
    """
    Handler for ZIP and RAR archives.
    Extracts all CDR files and returns:
      - polars.DataFrame  if all inner files are CSV/TXT (concatenated vertically)
      - dict              if inner files include Excel sheets or mixed types

    The archive format (.zip/.rar) is determined from `filename` passed at
    construction time — NOT from the file_obj passed to read(), because a
    BytesIO object carries no .name and would otherwise fall back to 'archive.bin',
    causing extract_files_from_archive to skip all entries.
    """

    def __init__(self, filename: str):
        # Store the original filename so read() always knows the archive format.
        self._filename = filename

    def read(self, file_obj):
        if isinstance(file_obj, (bytes, bytearray)):
            archive_bytes = file_obj
        else:
            file_obj.seek(0)
            archive_bytes = file_obj.read()

        # Use the filename stored at construction — never rely on file_obj.name
        archive_name = self._filename
        logger.debug(f"[ArchiveMemoryHandler] reading archive as '{archive_name}'")

        results = {}

        for inner_name, inner_bytes in extract_files_from_archive(archive_bytes, archive_name):
            inner_obj = io.BytesIO(inner_bytes)
            try:
                inner_handler = file_handler_factory_memory(inner_obj, inner_name)
                data = FileHandler(inner_handler).read_file(inner_obj)
                results[inner_name] = data
            except Exception as e:
                logger.error(f"[ARCHIVE HANDLER] Could not process '{inner_name}': {e}")

        if not results:
            raise ValueError("No valid CDR files found inside archive")

        all_dataframes = all(isinstance(v, pl.DataFrame) for v in results.values())
        if all_dataframes:
            dfs = list(results.values())
            return dfs[0] if len(dfs) == 1 else pl.concat(dfs, how="diagonal")

        return results


# ============================================================
# File-type Detection Helpers
# ============================================================

def detect_file_type(file_obj) -> str:
    """Detect CSV vs Excel by magic bytes."""
    try:
        file_obj.seek(0)
        header = file_obj.read(8)
        file_obj.seek(0)
        if header[:2] == b'PK':
            return 'xlsx'
        elif header[:4] == b'\xD0\xCF\x11\xE0':
            return 'xls'
        return 'csv'
    except Exception:
        return 'csv'


def detect_archive_type(file_obj) -> str:
    """Detect ZIP vs RAR by magic bytes."""
    try:
        file_obj.seek(0)
        header = file_obj.read(8)
        file_obj.seek(0)
        if header[:2] == b'PK':
            return 'zip'
        elif header[:7] == b'Rar!\x1a\x07':
            return 'rar'
        return 'unknown'
    except Exception:
        return 'unknown'


# ============================================================
# Memory-based Factory  (primary API)
# ============================================================

def file_handler_factory_memory(file_obj, filename: str) -> FileStrategy:
    """
    Return the appropriate FileStrategy for file_obj.

    Routing table:
      .csv / .txt          → CSVMemoryHandler
      .xls / .xlsx         → ExcelMemoryHandler  (CSV-content fallback)
      .zip / .rar          → ArchiveMemoryHandler
      .csv.xls / .csv.xlsx → CSVMemoryHandler     (misleading extension)
    """
    filename_lower = filename.lower()

    # Misleading double-extension
    if filename_lower.endswith('.csv.xls') or filename_lower.endswith('.csv.xlsx'):
        return CSVMemoryHandler()

    ext = os.path.splitext(filename_lower)[1]

    if ext in ('.zip', '.rar'):
        return ArchiveMemoryHandler(filename)

    if ext in ('.csv', '.txt'):
        return CSVMemoryHandler()

    if ext in ('.xls', '.xlsx'):
        file_type = detect_file_type(file_obj)
        return CSVMemoryHandler() if file_type == 'csv' else ExcelMemoryHandler()

    raise ValueError(f"Unsupported file type: '{ext}' in filename '{filename}'")


# ============================================================
# Disk-based Handlers  (backward compatibility)
# ============================================================

def get_path(relative_path: str) -> str:
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_dir, relative_path)


class CSVFileHandler(FileStrategy):

    def _detect_encoding(self, filepath: str) -> str:
        try:
            with open(filepath, 'rb') as f:
                rawdata = f.read(100000)
            return chardet.detect(rawdata).get('encoding', 'utf-8')
        except Exception:
            return 'utf-8'

    def _read_single_file(self, filepath: str) -> pl.DataFrame:
        full_path = filepath if os.path.isabs(filepath) else get_path(filepath)

        if not os.path.exists(full_path):
            raise FileNotFoundError(f"File not found: {full_path}")
        if os.path.getsize(full_path) == 0:
            raise ValueError("File is empty")

        encoding = self._detect_encoding(full_path)

        try:
            with open(full_path, "r", encoding=encoding, errors="ignore", newline="") as f:
                rows = list(csv.reader(f))

            if not rows:
                raise ValueError("CSV file has no data")

            max_cols = max(len(r) for r in rows)
            data_dict = {f"col_{i}": [] for i in range(max_cols)}
            for row in rows:
                for i in range(max_cols):
                    data_dict[f"col_{i}"].append(str(row[i]) if i < len(row) else "")

            return pl.DataFrame(
                data_dict,
                schema_overrides={col: pl.Utf8 for col in data_dict}
            )
        except Exception as e:
            raise ValueError(f"Error reading CSV file: {e}") from e

    def read(self, filepath):
        if isinstance(filepath, (list, tuple)):
            return pl.concat([self._read_single_file(f) for f in filepath], how="vertical")
        return self._read_single_file(filepath)


class ExcelHandler(FileStrategy):

    def read(self, filepath: str) -> dict:
        full_path = filepath if (os.path.isabs(filepath) and os.path.exists(filepath)) else get_path(filepath)

        try:
            wb = load_workbook(full_path, data_only=True)
            sheets = {}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                data = [tuple(cell.value for cell in row) for row in ws.iter_rows()]
                if not data:
                    continue

                max_cols = max(len(row) for row in data)
                if max_cols == 0:
                    continue

                normalized_data = [
                    (list(row) + [None] * (max_cols - len(row)))[:max_cols]
                    for row in data
                ]

                header = normalized_data[0]
                rows   = normalized_data[1:]

                header_clean = []
                seen_headers = {}
                for i, h in enumerate(header):
                    col_name = str(h).strip() if h is not None and str(h).strip() else f"col_{i}"
                    if col_name in seen_headers:
                        seen_headers[col_name] += 1
                        col_name = f"{col_name}_{seen_headers[col_name]}"
                    else:
                        seen_headers[col_name] = 0
                    header_clean.append(col_name)

                schema = {h: pl.Utf8 for h in header_clean}

                rows_clean = [
                    [
                        str(row[i]) if i < len(row) and row[i] is not None else ""
                        for i in range(max_cols)
                    ]
                    for row in rows
                ]

                if rows_clean:
                    sheets[sheet_name] = pl.DataFrame(rows_clean, schema=schema, orient="row")

            return sheets

        except Exception as e:
            raise ValueError(f"Cannot read Excel file: {filepath}") from e


class JSONStrategy(FileStrategy):

    def read(self, filepath: str):
        full_path = filepath if (os.path.isabs(filepath) and os.path.exists(filepath)) else get_path(filepath)
        with open(full_path, "r", encoding="utf-8") as f:
            return json.load(f)


# ============================================================
# Disk-based Factory  (backward compatibility)
# ============================================================

def file_handler_factory(file_path: str) -> FileStrategy:
    ext = os.path.splitext(file_path)[1].lower()

    if ext in ('.csv', '.txt'):
        return CSVFileHandler()
    elif ext in ('.xls', '.xlsx'):
        return ExcelHandler()
    elif ext == '.json':
        return JSONStrategy()
    else:
        raise ValueError(f"Unsupported file type: {ext}")